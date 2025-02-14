import streamlit as st
import pandas as pd
import plotly.express as px
import numpy as np
from streamlit_extras.metric_cards import style_metric_cards # beautify metric card with css
import plotly.graph_objects as go
import json
import openpyxl
from openpyxl import load_workbook
import altair as alt
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st. set_page_config(layout="wide")

#Chargement des bases de données
def read_excel_file(file):
    data = load_workbook(file)
    datas = data.active
    donnees = []
    for ligne in datas.iter_rows(values_only=True):
        donnees.append(list(ligne))
    en_tetes = donnees[0]
    donnees = donnees[1:]
    new_df = pd.DataFrame(donnees, columns=en_tetes)
    return new_df
df=read_excel_file("Base Pétrole finale.xlsx")
coord_géo= read_excel_file("Coordonnées géographiques Blocs.xlsx")
with open("GéoJson Blocs pétroliers.json") as f:
  counties = json.load(f)
st.sidebar.image('https://static.vecteezy.com/system/resources/thumbnails/010/248/729/original/national-emblem-coat-of-arms-or-symbol-of-ivory-coast-in-waving-flag-smooth-4k-seemless-loop-free-video.jpg', use_column_width='always')
st.sidebar.subheader("Cabinet du Ministre de l'Economie, du Plan et du Développement", divider="orange")

 #Définition des ordres de mois et des jours
order_of_months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
dic_month={1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}
order_of_days = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']
dic_day={0:"Lundi",1:"Mardi",2:"Mercredi",3:"Jeudi",4:"Vendredi",5:"Samedi",6:"Dimanche"}

def order_of_months_year(Année):
    order_of_months_year = []
    start_year = df[Année].min()
    end_year = df[Année].max()
    for year in range(start_year, end_year + 1):
        for month in order_of_months:
            order_of_months_year.append(f"{month} {year}")
    return order_of_months_year

def transf_df(df):
    for date in ["Date de signature du 1er CPP", "Date de la 2ème signature du CPP","Date de fin de validité d'exploration 1", "Date de fin de validité d'exploration 2","Date de fin de validité exploitation 1"]:
        df[date]= pd.to_datetime(df[date],format="%d/%m/%Y %H:%M", errors='coerce')
        df[f"Mois {str(date).replace('Date','')}"] = df[date].dt.month
        df[f"Jour {str(date).replace('Date','')}"] = df[date].dt.day_of_week
        df[f"heure {str(date).replace('Date','')}"]=df[date].dt.hour
        df[f"Année {str(date).replace('Date','')}"]=df[date].dt.year
        df[f"Mois {str(date).replace('Date','')}"]=df[f"Mois {str(date).replace('Date','')}"].map(dic_month)
        df[f"Mois* {str(date).replace('Date','')}"] = pd.Categorical(df[f"Mois {str(date).replace('Date','')}"], categories=order_of_months, ordered=True)
        df[f"Jour {str(date).replace('Date','')}"]=df[f"Jour {str(date).replace('Date','')}"].map(dic_day)
        df[f"Jour* {str(date).replace('Date','')}"] = pd.Categorical(df[f"Jour {str(date).replace('Date','')}"], categories=order_of_days, ordered=True)
        #df[f"Mois de l'année / {str(date).replace('Date','')}"] = df[f"Mois* {str(date).replace('Date','')}"].astype("str") + ' ' + df[f"Année {str(date).replace('Date','')}"].astype("str")
        #df[f"Mois de l'année / {str(date).replace('Date','')}"] = pd.Categorical(df[f"Mois de l'année / {str(date).replace('Date','')}"], categories=order_of_months_year(f"Année {str(date).replace('Date','')}"), ordered=True)    
    return df
df=transf_df(df)
page_bg_img = f"""
    <style>
    [data-testid="stAppViewContainer"] > .main {{
    background-image: url(https://www.vudaf.com/wp-content/uploads/2021/09/petrole-en-cote-divoire.jpg);
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    background-attachment: no-fixed;
    height: 100vh;
    margin: 0;
    display: flex;

    
    }}
    [data-testid="stSidebar"] {{
        background-color: #000 !important;  /* Fond noir */
        border: 2px solid #f7a900 !important;  /* Bordure rouge */
        border-radius: 10px;  /* Coins arrondis */
        margin-top: 0 px;  /* Ajuster la position vers le haut */
        position: relative;
        z-index: 1;  /* S'assurer que la barre latérale est au-dessus du contenu */
        padding: 10px;
    }}

    [data-testid="stHeader"] {{
    background: rgba(0, 0, 0, 0);
    color: white;
    }}

    [data-testid="stToolbar"] {{
    right: 2rem;
    }}
    </style>
    """

st.markdown(page_bg_img, unsafe_allow_html=True)
st.markdown('<div style="text-align:center;width:100%;"><h1 style="color:white;background-color:black;border:red;border-style:solid;border-radius:5px;">TABLEAU DE BORD DU SECTEUR PETROLIER AMONT IVOIRIEN </h1></div>', unsafe_allow_html=True)
st.write("")



def make_donut(input_response, input_text, input_color):
  if input_color == 'blue':
      chart_color = ['#29b5e8', '#155F7A']
  if input_color == 'green':
      chart_color = ['#27AE60', '#12783D']
  if input_color == 'orange':
      chart_color = ['#F39C12', '#875A12']
  if input_color == 'red':
      chart_color = ['#E74C3C', '#781F16']
    
  source = pd.DataFrame({
      "Topic": ['', input_text],
      "% value": [100-input_response, input_response]
  })
  source_bg = pd.DataFrame({
      "Topic": ['', input_text],
      "% value": [100, 0]
  })
    
  plot = alt.Chart(source).mark_arc(innerRadius=45, cornerRadius=25).encode(
      theta="% value",
      color= alt.Color("Topic:N",
                      scale=alt.Scale(
                          #domain=['A', 'B'],
                          domain=[input_text, ''],
                          # range=['#29b5e8', '#155F7A']),  # 31333F
                          range=chart_color),
                      legend=None),
  ).properties(width=130, height=130)
    
  text = plot.mark_text(align='center', color="#29b5e8", font="Lato", fontSize=32, fontWeight=700, fontStyle="italic").encode(text=alt.value(f'{input_response} %'))
  plot_bg = alt.Chart(source_bg).mark_arc(innerRadius=45, cornerRadius=20).encode(
      theta="% value",
      color= alt.Color("Topic:N",
                      scale=alt.Scale(
                          # domain=['A', 'B'],
                          domain=[input_text, ''],
                          range=chart_color),  # 31333F
                      legend=None),
  ).properties(width=130, height=130)
  return plot_bg + plot + text




def map():
    fig_map = px.choropleth_mapbox(df, geojson=counties, locations='Blocs',featureidkey="properties.name", color='Statut du bloc',
                            color_continuous_scale="Viridis",
                            range_color=(0, 12),
                            mapbox_style="open-street-map",
                            zoom=6,center = {"lat": 5.8, "lon": -5.61},
                            opacity=0.5,
                            labels={'Statut du bloc':'Statut du bloc'},
                            custom_data= [df["Opérateur le plus récent"], df['Superfice (en Km²)'], df["Type de profondeur"], df["Prod. Pétrole 2022 Bbls"], df['Prod Gaz N. 2022 MMSCF']]
                            )
    fig_map.update_traces(text=df['Blocs'], hovertemplate='<b>Opérateur le plus récent</b>: %{customdata[0]}<br>'
                                                            '<b>Superficie</b>: %{customdata[1]} Km2<br>'
                                                            '<b>Type de profondeur</b>: %{customdata[2]}<br>'
                                                            '<b>Production de pétrole du Bloc en 2022</b>: %{customdata[3]} Bbls<br>'
                                                            '<b>Production de pétrole du Bloc en 2022</b>: %{customdata[4]} MMSCF',
                                                            hoverlabel=dict(font=dict(size=10, color='white')),)
    fig_map.update_layout(margin={"r":0,"t":0,"l":0,"b":0})
    fig_map.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',})
    return st.plotly_chart(fig_map,use_container_width=True)


year=st.sidebar.select_slider("sélectionnez votre période d'analyse", options=[2018,2019,2020,2021,2022],value=(2021,2022))
st.subheader(f"Productions en {year[1]} par rapport à {year[0]}", divider="rainbow")

col = st.columns((5,5), gap='medium')
with col[0]:
    l=(df[f"Prod. Pétrole {year[1]} Bbls"].sum()-df[f"Prod. Pétrole {year[0]} Bbls"].sum())/1000
    st.metric(label=f"Production totale de barils de pétrole en {year[1]} (Bbls)", value=f"{((df[f'Prod. Pétrole {year[1]} Bbls'].sum())/1000000).round(2)} M Bbls", delta=f"{l} K Bbls")
    
with col[1]:
    p=(df[f"Prod Gaz N. {year[1]} MMSCF"].sum()-df[f"Prod Gaz N. {year[0]} MMSCF"].sum())
    st.metric(label=f"Production totale de barils de Gaz naturel en {year[1]}", value=f"{((df[f'Prod Gaz N. {year[1]} MMSCF'].sum())/1000).round(2)} K MMSCF", delta= f"{p} MMSCF")

st.subheader("Proportion de :", divider="rainbow") 
col2=st.columns((2.5,2.5,2.5,2.5), gap='medium')
  
with col2[0]:
    st.write("blocs en production")
    st.altair_chart(make_donut((df[df["Statut du bloc"]=="En activité - production"].shape[0]/df.shape[0])*100, 'Proportion de blocs en production', 'green'))
with col2[1]:    
    st.write("blocs en exploration")
    st.altair_chart(make_donut((df[df["Statut du bloc"]=="En activité - exploration"].shape[0]/df.shape[0])*100, 'Proportion de blocs exploration', 'blue'))
with col2[2]:
    st.write("blocs en négociation")
    st.altair_chart(make_donut(np.round(((df[df["Statut du bloc"]=="En activité - négociation"].shape[0]/df.shape[0])*100)), 'Proportion de blocs en négociation', 'orange'))
with col2[3]:
    st.write("blocs libres")
    st.altair_chart(make_donut((df[df["Statut du bloc"]=="Libre"].shape[0]/df.shape[0])*100, 'Proportion de blocs libres', 'red'))

style_metric_cards(background_color='#0c0c0c',border_left_color="#f7a900",box_shadow=True)


map()

st.header("Base de données personnalisée",divider="rainbow" )
def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    modify = st.checkbox("AJOUTEZ UN FILTRE")
    

    if not modify:
        return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Choisissez les variables que vous souhaitez utiliser comme filtre", df.columns)
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            # Treat columns with < 10 unique values as categorical
            int_columns = df.select_dtypes(include="int").columns
            float_columns = df.select_dtypes(include="float").columns

            if is_numeric_dtype(df[column]) :
                _min = int(df[column].min())
                _max = int(df[column].max())
                user_num_input = right.slider(
                    f"Valeurs de {column}",
                    min_value=_min,
                    max_value=_max,
                    value=(_min, _max),
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Valeur de {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            elif is_categorical_dtype(df[column]) or df[column].unique().shape[0]<100:
                arr=df[column].unique()
                user_cat_input = right.multiselect(
                    f"Valueur de {column}",
                    arr
                    ,
                    default=list(arr),
                )
                df = df[df[column].isin(user_cat_input)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].astype(str).str.contains(user_text_input)]

    return df
    
df_perso=filter_dataframe(df)
st.dataframe(df_perso)

colors = px.colors.sequential.Rainbow_r
colors.extend(px.colors.sequential.Agsunset)
colors.extend(px.colors.sequential.Aggrnyl)

# SECTION GRAPHIQUE
st.header("Analyses graphiques", divider="rainbow")

#Analyse univariée
st.subheader("Analyse graphique avec une seule variable")
# Histogramme et Camembert sur la même ligne
cam, hist = st.columns(2,gap='medium')

with cam:
    st.subheader("CAMEMBERT")
    selected_categorical_variable_p = st.selectbox("***Sélectionnez une variable catégorielle pour le camembert***", ['Type de profondeur', 'Opérateur1',
    'Patenaires (hors PETROCI)', 'Opérateur CPP 2',
    'Patenaires CPP 2 (hors PETROCI)', 'Opérateur CPP 3',
    'Patenaires CPP 3 (hors PETROCI)', 'Statut du bloc'], index=1)
    category_counts = df[selected_categorical_variable_p].value_counts()
    fig_pie = px.pie(names=category_counts.index, values=category_counts.values, title=f"Répartition de la variable {selected_categorical_variable_p}",color_discrete_sequence=colors)
    fig_pie.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.25)
    st.plotly_chart(fig_pie, use_container_width=True)

with hist:
    st.subheader("HISTOGRAMME")
    selected_categorical_variable = st.selectbox("***Sélectionnez la variable catégorielle pour l'histogramme***",['Type de profondeur', 'Opérateur1',
    'Patenaires (hors PETROCI)', 'Opérateur CPP 2',
    'Patenaires CPP 2 (hors PETROCI)', 'Opérateur CPP 3',
    'Patenaires CPP 3 (hors PETROCI)', 'Statut du bloc','Mois  de signature du 1er CPP', 'Année  de signature du 1er CPP', 'Année  de la 2ème signature du CPP', 'Mois  de la 2ème signature du CPP', "Mois  de fin de validité d'exploration 1", "Année  de fin de validité d'exploration 1", "Année  de fin de validité d'exploration 2", "Mois  de fin de validité d'exploration 2",'Mois  de fin de validité exploitation 1','Année  de fin de validité exploitation 1'], index=1)
    fig_histogram = px.histogram(df, x=df[selected_categorical_variable], color=df[selected_categorical_variable],title=f"Histogramme de {selected_categorical_variable}",color_discrete_sequence=colors)
    fig_histogram.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.35)
    fig_histogram.update_traces( textfont_color='rgba(255, 255, 255, 1)')
    if selected_categorical_variable in ['Mois  de signature du 1er CPP',  'Mois  de la 2ème signature du CPP', "Mois  de fin de validité d'exploration 1","Mois  de fin de validité d'exploration 2",'Mois  de fin de validité exploitation 1']:
        fig_histogram.update_xaxes(categoryorder='array', categoryarray=order_of_months)
    fig_histogram.update_xaxes(showticklabels=False)
    st.plotly_chart(fig_histogram,use_container_width=True)


# Section des analyses croisées
st.subheader("Analyse graphique avec deux variables croisées")



quant,qual=st.columns(2,gap='medium')


with quant:
    st.subheader("ANALYSE CROISEE ENTRE VARIABLES NUMERIQUES")
    int_columns = df.select_dtypes(include="int").columns
    float_columns = df.select_dtypes(include="float").columns
    selected_variable_3 = st.selectbox("***Variable 1***", int_columns.union(float_columns))
    selected_variable_4 = st.selectbox("***Variable 2***",int_columns.union(float_columns),index=2)
    fig_scatter_matrix = px.scatter(df, x=selected_variable_3, y=selected_variable_4)
    fig_scatter_matrix.update_layout(title=f'Nuage de points entre {selected_variable_3} et {selected_variable_4}')
    fig_scatter_matrix.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.15)
    st.plotly_chart(fig_scatter_matrix, use_container_width=True)


with qual:
    #Type de l'histogramme croisé
    def barmode_selected(t):
        if t =='empilé':
            a='relative'  
        else: 
            a='group'
        return a
    
    st.subheader("ANALYSE CROISEE ENTRE VARIABLES CATEGORIELLES")
    selected_variable_1 = st.selectbox("***Variable 1***", ['Type de profondeur', 'Opérateur1',
    'Patenaires (hors PETROCI)', 'Opérateur CPP 2',
    'Patenaires CPP 2 (hors PETROCI)', 'Opérateur CPP 3',
    'Patenaires CPP 3 (hors PETROCI)', 'Statut du bloc','Mois  de signature du 1er CPP', 'Année  de signature du 1er CPP', 'Année  de la 2ème signature du CPP', 'Mois  de la 2ème signature du CPP', "Mois  de fin de validité d'exploration 1", "Année  de fin de validité d'exploration 1", "Année  de fin de validité d'exploration 2", "Mois  de fin de validité d'exploration 2",'Mois  de fin de validité exploitation 1','Année  de fin de validité exploitation 1'], index=1)
    selected_variable_2 = st.selectbox("***Variable 2***", ['Type de profondeur', 'Opérateur1',
    'Patenaires (hors PETROCI)', 'Opérateur CPP 2',
    'Patenaires CPP 2 (hors PETROCI)', 'Opérateur CPP 3',
    'Patenaires CPP 3 (hors PETROCI)', 'Statut du bloc','Mois  de signature du 1er CPP', 'Année  de signature du 1er CPP', 'Année  de la 2ème signature du CPP', 
    'Mois  de la 2ème signature du CPP', "Mois  de fin de validité d'exploration 1", "Année  de fin de validité d'exploration 1", 
    "Année  de fin de validité d'exploration 2", "Mois  de fin de validité d'exploration 2",'Mois  de fin de validité exploitation 1',
    'Année  de fin de validité exploitation 1'],index=2)
    st.sidebar.write(" ")
    st.sidebar.write(" ")
    st.sidebar.subheader("PARAMETRES DES GRAPHIQUES")
    type_graph=st.sidebar.radio("***:grey[Choisissez le type d'histogramme croisé]***", ['empilé','étalé'])
    if selected_variable_2 in [f"Prod Gaz N. {i} MMSCF" for i in range(2018,2023)] or [f"Prod. Pétrole {i} Bbls" for i in range(2018,2023)]:
        fig_croisé = px.bar(df.groupby(selected_variable_1)[selected_variable_2].sum().reset_index(), x=selected_variable_1,y=selected_variable_2, color=selected_variable_2,barmode=barmode_selected(type_graph),color_continuous_scale=['red', 'yellow', 'green'],range_color=[0, 5])
    else:
        fig_croisé = px.bar(df, x=selected_variable_1, color=selected_variable_2,barmode=barmode_selected(type_graph),color_discrete_sequence= colors)
        m=['Mois  de signature du 1er CPP', 'Mois  de la 2ème signature du CPP', "Mois  de fin de validité d'exploration 1", "Mois  de fin de validité d'exploration 2",'Mois  de fin de validité exploitation 1']
        if selected_variable_1 in m or selected_variable_2 in m:
            fig_croisé.update_xaxes(categoryorder='array', categoryarray=order_of_months)
    fig_croisé.update_layout(title=f'Graphique en barres groupées - {selected_variable_1 } vs {selected_variable_2 }')
    fig_croisé.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.20)

    st.plotly_chart(fig_croisé,use_container_width=True)


